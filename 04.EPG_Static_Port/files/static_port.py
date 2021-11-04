from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

class static_port(object):

    def __init__(self):
        self.dict_data      = {}
        self.in_excel_file  = "static_ports.xlsx"
        self.out_tfvar_file = "../static_ports.auto.tfvars"
        self.template_file  = "static_ports.auto.tfvars.template"


    def main(self):
        print("##########################################")
        print("#                                        #")
        print("#  xlsx-to-tfvars for ACI Static Port    #")
        print("#  For Demo                              #")
        print("#  version: 0.1                          #")
        print("#                                        #")
        print("##########################################")
        print("File readed: %s" % self.in_excel_file)

    def read(self):
        workbook = load_workbook(filename=self.in_excel_file)
        self.keys = workbook.sheetnames
        for sheet in workbook:
            data = sheet.values
            cols = next(data)[1:]
            self.dict_data = list(data)
    
    def write(self):
        env = Environment(loader=FileSystemLoader('./'))
        template1 = env.get_template(self.template_file)
        output = []
        output.append(template1.render(data=self.dict_data))
        with open(self.out_tfvar_file, mode='w+') as f:
            f.writelines(output)
        print("File created: %s" % self.out_tfvar_file)

if __name__ == "__main__":
    func = static_port()
    func.main()
    func.read()
    func.write()

